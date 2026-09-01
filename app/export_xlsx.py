# -*- coding: utf-8 -*-
"""
Генерация Excel-отчёта "Регионы × ТК" в светлой цветовой гамме,
с двумя рекомендациями лучшего ТК для каждого города (и региона):
  1) Быстрее и надёжнее  — % выкупа + срок доставки
  2) Выгоднее по деньгам — % выкупа + тариф (отправка + возврат)
Финрез считается по формуле: ср. чек − тариф на доставку (с учётом возврата).

Формулы полностью повторяют те, что используются в веб-интерфейсе
(см. calcCityRecommendations в templates/index.html), чтобы цифры совпадали.
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TK_ORDER = ['Почта России', '5Post', 'СДЭК', 'Курьер (свой)', 'Курьер (стационар)']
SEGS_ORDER = ['до 20 тыс.', '20–50 тыс.', '50–100 тыс.', '100 тыс.+']

# ── Светлая палитра (пастельные тона вместо тёмной темы приложения) ──────
TITLE_BG   = 'E3EAF6'; TITLE_FG   = '0F172A'
REGION_BG  = 'EAEFF8'; REGION_FG  = '0F172A'
SEGMENT_BG = 'F3F7FF'; SEGMENT_FG = '1D4ED8'
CITY_BG    = 'FFFFFF'; CITY_FG    = '334155'
POS_FG = '059669'
NEG_FG = 'DC2626'
SOFT_FG = '94A3B8'
BORDER_CLR = 'E5EAF1'

TK_COLORS = {
    'Почта России':        ('FBEBD0', '92400E'),
    '5Post':                ('D6F3E6', '065F46'),
    'СДЭК':                 ('DEE7FB', '1E3A8A'),
    'Курьер (свой)':        ('EAE1FA', '5B21B6'),
    'Курьер (стационар)':   ('FCDFE4', '9F1239'),
}
REC1_COLOR = ('FFF1BE', '854D0E')  # ⚡ быстрее и надёжнее
REC2_COLOR = ('D6F5EF', '0F766E')  # 💰 выгоднее по деньгам

thin = Side(style='thin', color=BORDER_CLR)
CELL_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def _calc_recommendations(row):
    """Повторяет логику calcCityRecommendations() из фронтенда."""
    tkd = row.get('tk_breakdown') or {}
    avg_check = row.get('avg_check') or 0
    candidates = []
    for tk, t in tkd.items():
        total = t.get('total') or 0
        if total < 3:
            continue
        dr = t.get('delivery_rate') or 0
        days = t.get('days')
        rr = t.get('return_rate') or 0
        tarif_base = t.get('avg_tarif') if t.get('avg_tarif') is not None else t.get('avg_cost')
        tarif_full = tarif_base * (1 + rr / 100) if tarif_base is not None else None
        finrez = round(avg_check - tarif_full) if tarif_full is not None else None
        candidates.append({
            'tk': tk, 'dr': dr, 'days': days, 'tarif_full': tarif_full, 'finrez': finrez,
        })

    if not candidates:
        return None, None

    def score_speed(c):
        sv = c['dr'] or 0
        ss = max(0, (1 - (c['days'] if c['days'] is not None else 20) / 30) * 100)
        return sv * 0.6 + ss * 0.4

    with_days = [c for c in candidates if c['days'] is not None]
    pool1 = with_days or candidates
    best1 = max(pool1, key=score_speed)

    with_tarif = [c for c in candidates if c['tarif_full'] is not None]
    pool2 = with_tarif or candidates
    tarifs = [c['tarif_full'] for c in pool2 if c['tarif_full'] is not None]
    min_t, max_t = (min(tarifs), max(tarifs)) if tarifs else (0, 0)

    def score_cost(c):
        if c['tarif_full'] is None:
            return (c['dr'] or 0) * 0.6
        cs = (1 - (c['tarif_full'] - min_t) / (max_t - min_t)) * 100 if max_t > min_t else 100
        return (c['dr'] or 0) * 0.6 + cs * 0.4

    best2 = max(pool2, key=score_cost)
    return best1, best2


def _fill(hexcolor):
    return PatternFill('solid', fgColor=hexcolor)


def _write_header(ws):
    n_tk = len(TK_ORDER)
    last_col = 4 + n_tk * 4 + 8  # A-D + TK blocks(4 each) + 2 rec blocks(4 each)

    # Строка 1 — заголовок
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    c = ws.cell(1, 1, 'Аналитика доставки · Регионы × ТК (с рекомендациями по ТК)')
    c.font = Font(bold=True, size=13, color=TITLE_FG)
    c.fill = _fill(TITLE_BG)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[1].height = 30

    # Строка 2 — групповые заголовки
    col = 5
    for tk in TK_ORDER:
        bg, fg = TK_COLORS[tk]
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 3)
        c = ws.cell(2, col, tk)
        c.font = Font(bold=True, color=fg, size=10)
        c.fill = _fill(bg)
        c.alignment = Alignment(horizontal='center')
        col += 4

    for label, (bg, fg) in [('⚡ Быстрее и надёжнее', REC1_COLOR), ('💰 Выгоднее по деньгам', REC2_COLOR)]:
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 3)
        c = ws.cell(2, col, label)
        c.font = Font(bold=True, color=fg, size=10)
        c.fill = _fill(bg)
        c.alignment = Alignment(horizontal='center')
        col += 4
    ws.row_dimensions[2].height = 20

    # Строка 3 — подписи столбцов
    headers = ['Регион', 'Сегмент / Город', 'Заказов', '% выкупа']
    for tk in TK_ORDER:
        headers += ['% выкупа', 'Срок, дн.', 'Тариф, ₽', 'Заказов']
    headers += ['ТК', '% выкупа', 'Срок, дн.', 'Финрез, ₽']
    headers += ['ТК', '% выкупа', 'Тариф, ₽', 'Финрез, ₽']

    for i, h in enumerate(headers, start=1):
        c = ws.cell(3, i, h)
        c.font = Font(bold=True, size=9, color='475569')
        c.fill = _fill('F8FAFC')
        c.alignment = Alignment(horizontal='center', wrap_text=True)
        c.border = CELL_BORDER
    ws.row_dimensions[3].height = 30

    # Ширины колонок
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    for i in range(5, last_col + 1):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = 10.5

    ws.freeze_panes = 'A4'
    return last_col


def _pct_cell(ws, r, col, value):
    c = ws.cell(r, col)
    if value is None:
        return c
    c.value = round(value, 1) / 100
    c.number_format = '0.0%'
    c.font = Font(color=POS_FG if value >= 50 else NEG_FG, bold=True, size=10)
    c.alignment = Alignment(horizontal='center')
    return c


def _num_cell(ws, r, col, value, fmt='#,##0', center=True, color=None, bold=False):
    c = ws.cell(r, col)
    if value is None:
        return c
    c.value = value
    c.number_format = fmt
    c.font = Font(size=10, color=color, bold=bold)
    if center:
        c.alignment = Alignment(horizontal='center')
    return c


def _write_tk_block(ws, r, col, t):
    if not t:
        col += 4
        return col
    _pct_cell(ws, r, col, t.get('delivery_rate'))
    _num_cell(ws, r, col + 1, t.get('days'), fmt='0.0')
    tarif = t.get('avg_tarif') if t.get('avg_tarif') is not None else t.get('avg_cost')
    _num_cell(ws, r, col + 2, round(tarif) if tarif is not None else None, fmt='#,##0" ₽"')
    _num_cell(ws, r, col + 3, t.get('total'), fmt='#,##0')
    return col + 4


def _write_reco_block(ws, r, col, best, metric_key, metric_fmt):
    if not best:
        return col + 4
    ws.cell(r, col, best['tk']).font = Font(size=10, bold=True)
    ws.cell(r, col).alignment = Alignment(horizontal='center')
    _pct_cell(ws, r, col + 1, best['dr'])
    metric_val = best.get(metric_key)
    if metric_key == 'tarif_full' and metric_val is not None:
        metric_val = round(metric_val)
    _num_cell(ws, r, col + 2, metric_val, fmt=metric_fmt)
    fin = best.get('finrez')
    _num_cell(ws, r, col + 3, fin, fmt='#,##0" ₽"',
              color=(POS_FG if (fin or 0) > 0 else NEG_FG), bold=True)
    return col + 4


def _row_style(ws, r, last_col, bg, name_col, name_font):
    for i in range(1, last_col + 1):
        c = ws.cell(r, i)
        c.fill = _fill(bg)
        c.border = CELL_BORDER
    ws.cell(r, name_col).font = name_font


def build_regions_xlsx(rows: list) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Регионы × ТК'
    last_col = _write_header(ws)

    r = 4
    for region in rows:
        # ── строка региона ──
        ws.cell(r, 1, region.get('region', '')).font = Font(bold=True, color=REGION_FG, size=11)
        ws.cell(r, 1).alignment = Alignment(horizontal='left', indent=1)
        _num_cell(ws, r, 3, region.get('total'), fmt='#,##0', bold=True)
        _pct_cell(ws, r, 4, region.get('delivery_rate'))

        col = 5
        tkd = region.get('tk_breakdown') or {}
        for tk in TK_ORDER:
            col = _write_tk_block(ws, r, col, tkd.get(tk))

        best1, best2 = _calc_recommendations(region)
        col = _write_reco_block(ws, r, col, best1, 'days', '0.0')
        col = _write_reco_block(ws, r, col, best2, 'tarif_full', '#,##0" ₽"')

        _row_style(ws, r, last_col, REGION_BG, 1, Font(bold=True, color=REGION_FG, size=11))
        ws.row_dimensions[r].height = 20
        r += 1

        # ── строки сегментов + городов ──
        cities = region.get('cities') or []
        by_seg = {}
        for c in cities:
            by_seg.setdefault(c.get('segment') or '—', []).append(c)

        for seg in SEGS_ORDER:
            seg_cities = by_seg.get(seg) or []
            if not seg_cities:
                continue

            seg_total = sum(c.get('total') or 0 for c in seg_cities)
            seg_del = sum(c.get('delivered') or 0 for c in seg_cities)
            seg_pct = round(seg_del / seg_total * 100, 1) if seg_total else 0

            ws.cell(r, 2, '📦 ' + seg).alignment = Alignment(horizontal='left', indent=1)
            _num_cell(ws, r, 3, seg_total, fmt='#,##0', bold=True)
            _pct_cell(ws, r, 4, seg_pct)
            _row_style(ws, r, last_col, SEGMENT_BG, 2, Font(bold=True, color=SEGMENT_FG, size=10))
            r += 1

            for c in seg_cities:
                ws.cell(r, 2, '   ' + str(c.get('city', ''))).alignment = Alignment(horizontal='left', indent=1)
                _num_cell(ws, r, 3, c.get('total'), fmt='#,##0')
                _pct_cell(ws, r, 4, c.get('delivery_rate'))

                col = 5
                ctk = c.get('tk_breakdown') or {}
                for tk in TK_ORDER:
                    col = _write_tk_block(ws, r, col, ctk.get(tk))

                cb1, cb2 = _calc_recommendations(c)
                col = _write_reco_block(ws, r, col, cb1, 'days', '0.0')
                col = _write_reco_block(ws, r, col, cb2, 'tarif_full', '#,##0" ₽"')

                _row_style(ws, r, last_col, CITY_BG, 2, Font(color=CITY_FG, size=10))
                r += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
